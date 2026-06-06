from decimal import Decimal
import csv
from typing import Callable

import openpyxl

from . import storage
from .ast import (
    DropTable, CreateTable, InsertValues, DeleteStatement, SelectStatement, UpdateStatement, ImportStatement,
    AggFunc, AggregateColumn,
)
from .evaluator import evaluate_where
from .exceptions import ColumnNotFoundError, TableNotFoundError, EngineError
from .types import Types, RowDict, DataValue

# ---------------------------------------------------------------------------
# Type map: parser type strings → Types enum
# ---------------------------------------------------------------------------

_TYPE_MAP = {
    'STRING': Types.STRING,
    'NUMBER': Types.NUMBER,
}


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

def execute_drop(statement: DropTable) -> None:
    storage.drop_table(statement.table_name)


def execute_create(statement: CreateTable) -> None:
    defs = {col_name: _TYPE_MAP[type_str] for col_name, type_str in statement.columns}
    storage.create_table(statement.table_name, defs)


# ---------------------------------------------------------------------------
# DML
# ---------------------------------------------------------------------------

def execute_delete(statement: DeleteStatement) -> int:
    table = storage.get_table(statement.table_name)
    if table is None:
        raise TableNotFoundError(f'Table does not exist: {statement.table_name}')

    where_func = None
    if statement.where_clause is not None:
        where_func = evaluate_where(statement.where_clause.expression, table.column_types())

    return table.delete(where=where_func)


def execute_insert(statement: InsertValues) -> None:
    table = storage.get_table(statement.table_name)
    if table is None:
        raise TableNotFoundError(f'Table does not exist: {statement.table_name}')

    all_columns = table.column_names()

    # Determine target columns
    if statement.columns is None:
        target_columns = all_columns
        if len(target_columns) != len(statement.values[0]):
            raise ValueError(f'Column count mismatch: expected {len(target_columns)}, '
                             f'got {len(statement.values[0])}')
    else:
        for col in statement.columns:
            if col not in all_columns:
                raise ColumnNotFoundError(f'Column does not exist: {col}')
        target_columns = statement.columns

    for value_row in statement.values:
        if len(target_columns) != len(value_row):
            raise ValueError(f'Column count mismatch: expected {len(target_columns)}, '
                             f'got {len(value_row)}')

        row_data: RowDict = {}
        for col, val in zip(target_columns, value_row):
            row_data[col] = val

        for col in all_columns:
            if col not in row_data:
                row_data[col] = None

        table.insert_values(row_data)


# ---------------------------------------------------------------------------
# Aggregate helpers
# ---------------------------------------------------------------------------

def _make_agg_func(agg_col: AggregateColumn):
    """Build an aggregation function for a single AggregateColumn."""
    func = agg_col.func
    col = agg_col.column

    match func:
        case AggFunc.COUNT:
            if col is None:
                return lambda group: Decimal(len(group))
            return lambda group, _c=col: Decimal(sum(1 for row in group if row.get(_c) is not None))

        case AggFunc.SUM:
            def _sum(group, _c=col):
                vals = [row[_c] for row in group if row.get(_c) is not None]
                return sum(vals, Decimal(0)) if vals else None
            return _sum

        case AggFunc.AVG:
            def _avg(group, _c=col):
                vals = [row[_c] for row in group if row.get(_c) is not None]
                return sum(vals, Decimal(0)) / len(vals) if vals else None
            return _avg

        case AggFunc.MIN:
            def _min(group, _c=col):
                vals = [row[_c] for row in group if row.get(_c) is not None]
                return min(vals) if vals else None
            return _min

        case AggFunc.MAX:
            def _max(group, _c=col):
                vals = [row[_c] for row in group if row.get(_c) is not None]
                return max(vals) if vals else None
            return _max


def _group_rows(rows: list[RowDict], group_by: list[str] | None) -> dict[tuple, list[RowDict]]:
    """Group rows by given columns. Returns {key_tuple: [rows]}."""
    groups: dict[tuple, list[RowDict]] = {}
    if group_by:
        for row in rows:
            key = tuple(row[col] for col in group_by)
            groups.setdefault(key, []).append(row)
    else:
        groups = {(): rows}
    return groups


def _apply_aggregations(
    groups: dict[tuple, list[RowDict]],
    group_by: list[str] | None,
    aggregates: dict[str, Callable],
) -> list[RowDict]:
    """Apply aggregate functions to each group and return result rows."""
    result: list[RowDict] = []
    for key, group_rows in groups.items():
        row_result: RowDict = {}
        if group_by:
            for col, val in zip(group_by, key):
                row_result[col] = val
        for alias, agg_func in aggregates.items():
            row_result[alias] = agg_func(group_rows)
        result.append(row_result)
    return result


# ---------------------------------------------------------------------------
# SELECT
# ---------------------------------------------------------------------------

def execute_select(statement: SelectStatement) -> list[RowDict]:
    table = storage.get_table(statement.table_name)
    if table is None:
        raise TableNotFoundError(f'Table does not exist: {statement.table_name}')

    where_func = None
    if statement.where_clause is not None:
        where_func = evaluate_where(statement.where_clause.expression, table.column_types())

    # Non-aggregate path
    if not statement.aggregates:
        result = table.select(
            columns=statement.columns,
            where=where_func,
            order_by=statement.order_by,
            distinct=statement.distinct,
            limit=statement.limit,
            offset=statement.offset,
        )
    else:
        # Aggregate path
        rows = table.select(where=where_func)

        # Build aggregation functions
        aggregations = {agg_col.alias: _make_agg_func(agg_col) for agg_col in statement.aggregates}

        # Group and aggregate
        groups = _group_rows(rows, statement.group_by)
        result = _apply_aggregations(groups, statement.group_by, aggregations)

        # HAVING
        if statement.having is not None:
            having_defs: dict[str, Types] = {}
            if statement.group_by:
                for gcol in statement.group_by:
                    if gcol in table.column_types():
                        having_defs[gcol] = table.column_types()[gcol]
            for agg_col in statement.aggregates:
                having_defs[agg_col.alias] = Types.NUMBER
            having_func = evaluate_where(statement.having.expression, having_defs)
            result = [r for r in result if having_func(r)]

        # ORDER BY
        if statement.order_by:
            for col, desc in reversed(statement.order_by):
                null_rows = [r for r in result if r.get(col) is None]
                non_null_rows = [r for r in result if r.get(col) is not None]
                non_null_rows.sort(key=lambda r, _c=col: r[_c], reverse=desc)
                result = non_null_rows + null_rows

        # LIMIT / OFFSET
        if statement.offset > 0:
            result = result[statement.offset:]
        if statement.limit is not None:
            result = result[:statement.limit]

    # INTO OUTFILE (both paths)
    if statement.output_file is not None:
        all_cols = table.column_names()
        headers = statement.columns if statement.columns else all_cols
        return _write_output(result, headers, statement.output_file)

    return result


# ---------------------------------------------------------------------------
# UPDATE
# ---------------------------------------------------------------------------

def execute_update(statement: UpdateStatement) -> int:
    table = storage.get_table(statement.table_name)
    if table is None:
        raise TableNotFoundError(f'Table does not exist: {statement.table_name}')

    all_columns = table.column_names()
    for col, _ in statement.set_clauses:
        if col not in all_columns:
            raise ColumnNotFoundError(f'Column does not exist: {col}')

    where_func = None
    if statement.where_clause is not None:
        where_func = evaluate_where(statement.where_clause.expression, table.column_types())

    affected = len(table.select(where=where_func))
    values: RowDict = {col: val for col, val in statement.set_clauses}
    table.update(values, where=where_func)

    return affected


# ---------------------------------------------------------------------------
# EXPORT helper
# ---------------------------------------------------------------------------

def _write_output(rows: list[RowDict], headers: list[str], filepath: str) -> int:
    """Write query results to file. Format determined by extension."""
    ext = filepath.rsplit('.', 1)[-1].lower() if '.' in filepath else ''

    if ext == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append([_cell_value(row.get(h)) for h in headers])
        wb.save(filepath)
    else:
        delim = '\t' if ext == 'tsv' else ','
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=delim, quoting=csv.QUOTE_NONNUMERIC)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([_cell_value(row.get(h)) for h in headers])

    return len(rows)


def _cell_value(value: object) -> object:
    """Convert data value to a format suitable for export."""
    if value is None:
        return ''
    return value


# ---------------------------------------------------------------------------
# IMPORT
# ---------------------------------------------------------------------------

def _infer_types(rows: list[tuple], headers: list[str]) -> dict[str, Types]:
    """Scan first 100 rows to infer column types for each header."""
    sample = rows[:100]
    defs: dict[str, Types] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        is_numeric = all(
            row[i] is None or isinstance(row[i], (int, float))
            for row in sample
        )
        defs[h] = Types.NUMBER if is_numeric else Types.STRING
    return defs


def execute_import(statement: ImportStatement) -> int:
    """Import an Excel file into a new table."""
    try:
        wb = openpyxl.load_workbook(statement.file_path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise EngineError(f'File not found: {statement.file_path}')

    ws = wb.active
    raw_rows: list[tuple] = list(ws.iter_rows(values_only=True))
    wb.close()

    if not raw_rows:
        raise EngineError('Excel file is empty')

    headers = [str(cell) if cell is not None else '' for cell in raw_rows[0]]

    # Column definitions: user-specified or auto-inferred
    if statement.columns:
        col_defs = {name: Types(t) for name, t in statement.columns}
    else:
        col_defs = _infer_types(raw_rows[1:], headers)

    storage.create_table(statement.table_name, col_defs)
    table = storage.get_table(statement.table_name)
    if table is None:
        raise EngineError(f'Failed to create table: {statement.table_name}')

    count = 0
    for row in raw_rows[1:]:
        values: RowDict = {}
        for i, h in enumerate(headers):
            if h:
                values[h] = row[i] if i < len(row) else None
        table.insert_values(values)
        count += 1

    return count
